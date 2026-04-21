# -*- coding: utf-8 -*-
"""
超无穹AI平台培训课程 - Excel文档生成
基于openpyxl库
"""

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os

OUTPUT_DIR = r"D:\网站部署\超无穹项目\chaowuqiong-project\docs\培训课程\文档生成\excel"

def set_header_style(cell, color='4472C4'):
    cell.font = Font(name='微软雅黑', size=11, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
    cell.alignment = Alignment(horizontal='center', vertical='center')

def set_cell_style(cell, bold=False, align='left'):
    cell.font = Font(name='微软雅黑', size=10)
    cell.alignment = Alignment(horizontal=align, vertical='center', wrap_text=True)

def set_border(cell):
    thin = Side(style='thin')
    cell.border = Border(left=thin, right=thin, top=thin, bottom=thin)

def create_schedule_excel():
    """创建课程表Excel"""
    wb = Workbook()

    ws = wb.active
    ws.title = "课程总表"

    headers = ['周次', '阶段', '课时', '日期', '星期', '时间', '课程内容', '实践任务', '学习目标']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_header_style(cell)

    schedule_data = [
        ('第1周', '🌟 入门篇', 6, '', '周一至周六', '9:00-10:00', '计算机基础与开发环境', '安装开发软件、配置环境', '掌握计算机基本操作'),
        ('第1周', '🌟 入门篇', 6, '', '周一至周六', '10:30-11:30', '计算机基本操作', '使用文件管理系统', '熟悉操作系统'),
        ('第1周', '🌟 入门篇', 6, '', '周一至周六', '14:00-15:00', '认识互联网', '注册平台账号', '理解网络基础'),
        ('第1周', '🌟 入门篇', 6, '', '周一至周六', '15:30-16:30', '编程是什么', 'Hello World', '建立编程思维'),
        ('第1周', '🌟 入门篇', 6, '', '周一至周六', '17:00-18:00', '开发工具介绍', 'VS Code配置', '掌握开发工具'),
        ('第2周', '🌟 入门篇', 6, '', '周一至周六', '9:00-10:00', 'HTML初识', '个人介绍页面', '掌握HTML基础'),
        ('第2周', '🌟 入门篇', 6, '', '周一至周六', '10:30-11:30', 'CSS入门', '美化页面', '掌握CSS样式'),
        ('第2周', '🌟 入门篇', 6, '', '周一至周六', '14:00-15:00', '响应式设计基础', '自适应布局', '理解响应式原理'),
        ('第2周', '🌟 入门篇', 6, '', '周一至周六', '15:30-16:30', '浏览器开发者工具', '调试页面', '掌握调试技巧'),
        ('第2周', '🌟 入门篇', 6, '', '周一至周六', '17:00-18:00', '实践项目', '简历页面项目', '完成P1项目'),
        ('第3周', '🔧 基础篇', 6, '', '周一至周六', '9:00-10:00', 'JavaScript概述', 'JS引入HTML', '理解JS作用'),
        ('第3周', '🔧 基础篇', 6, '', '周一至周六', '10:30-11:30', '变量与数据类型', '数据类型练习', '掌握变量声明'),
        ('第3周', '🔧 基础篇', 6, '', '周一至周六', '14:00-15:00', '运算符与表达式', '计算器实现', '掌握运算'),
        ('第3周', '🔧 基础篇', 6, '', '周一至周六', '15:30-16:30', '条件语句', '登录验证', '掌握条件判断'),
        ('第3周', '🔧 基础篇', 6, '', '周一至周六', '17:00-18:00', '循环结构', '99乘法表', '掌握循环'),
        ('第4周', '🔧 基础篇', 6, '', '周一至周六', '9:00-10:00', '函数基础', '工具函数封装', '掌握函数'),
        ('第4周', '🔧 基础篇', 6, '', '周一至周六', '10:30-11:30', '数组基础', '通讯录管理', '掌握数组'),
        ('第4周', '🔧 基础篇', 6, '', '周一至周六', '14:00-15:00', '对象基础', '用户对象', '掌握对象'),
        ('第4周', '🔧 基础篇', 6, '', '周一至周六', '15:30-16:30', 'DOM操作', '留言板', '掌握DOM'),
        ('第4周', '🔧 基础篇', 6, '', '周一至周六', '17:00-18:00', '实践项目', 'Todo App', '完成P2项目'),
        ('第5周', '🔧 基础篇', 6, '', '周一至周六', '9:00-10:00', 'Node.js介绍', '第一个服务', '掌握Node基础'),
        ('第5周', '🔧 基础篇', 6, '', '周一至周六', '10:30-11:30', 'Express框架', '路由中间件', '掌握Express'),
        ('第5周', '🔧 基础篇', 6, '', '周一至周六', '14:00-15:00', '数据库概述', 'MySQL安装', '理解数据库'),
        ('第5周', '🔧 基础篇', 6, '', '周一至周六', '15:30-16:30', 'SQL基础', 'CRUD操作', '掌握SQL'),
        ('第5周', '🔧 基础篇', 6, '', '周一至周六', '17:00-18:00', '阶段实践', '用户系统', '完成P3项目'),
        ('第6周', '🚀 进阶篇', 6, '', '周一至周六', '9:00-10:00', 'React概述', '第一个组件', '理解React'),
        ('第6周', '🚀 进阶篇', 6, '', '周一至周六', '10:30-11:30', '组件与Props', 'Header组件', '掌握组件'),
        ('第6周', '🚀 进阶篇', 6, '', '周一至周六', '14:00-15:00', 'State与生命周期', '计数器', '掌握State'),
        ('第6周', '🚀 进阶篇', 6, '', '周一至周六', '15:30-16:30', '事件处理', '表单组件', '掌握事件'),
        ('第6周', '🚀 进阶篇', 6, '', '周一至周六', '17:00-18:00', 'Hooks深入', '自定义Hook', '掌握Hooks'),
        ('第7周', '🚀 进阶篇', 6, '', '周一至周六', '9:00-10:00', '项目结构分析', '源码阅读', '理解架构'),
        ('第7周', '🚀 进阶篇', 6, '', '周一至周六', '10:30-11:30', 'AI对话功能', '对话组件', '开发AI对话'),
        ('第7周', '🚀 进阶篇', 6, '', '周一至周六', '14:00-15:00', '角色管理', 'CRUD角色', '开发角色管理'),
        ('第7周', '🚀 进阶篇', 6, '', '周一至周六', '15:30-16:30', '会员系统前端', '会员中心', '开发会员'),
        ('第7周', '🚀 进阶篇', 6, '', '周一至周六', '17:00-18:00', '移动端适配', '响应式适配', '掌握适配'),
        ('第8周', '🚀 进阶篇', 6, '', '周一至周六', '9:00-10:00', '后端结构分析', '路由分析', '理解后端'),
        ('第8周', '🚀 进阶篇', 6, '', '周一至周六', '10:30-11:30', '认证中间件', 'JWT验证', '实现认证'),
        ('第8周', '🚀 进阶篇', 6, '', '周一至周六', '14:00-15:00', 'LLM配置模块', '多模型切换', '配置LLM'),
        ('第8周', '🚀 进阶篇', 6, '', '周一至周六', '15:30-16:30', '支付集成', '支付功能', '对接支付'),
        ('第8周', '🚀 进阶篇', 6, '', '周一至周六', '17:00-18:00', 'Redis缓存', 'Session管理', '实现缓存'),
        ('第9周', '💎 高阶篇', 6, '', '周一至周六', '9:00-10:00', 'TypeScript入门', 'TS改造', '掌握TypeScript'),
        ('第9周', '💎 高阶篇', 6, '', '周一至周六', '10:30-11:30', '状态管理', 'Redux/Zustand', '掌握状态管理'),
        ('第9周', '💎 高阶篇', 6, '', '周一至周六', '14:00-15:00', '前端路由', 'React Router', '掌握路由'),
        ('第9周', '💎 高阶篇', 6, '', '周一至周六', '15:30-16:30', '构建工具', 'Vite优化', '掌握构建'),
        ('第9周', '💎 高阶篇', 6, '', '周一至周六', '17:00-18:00', 'Git工作流', '分支管理', '掌握Git'),
        ('第10周', '💎 高阶篇', 6, '', '周一至周六', '9:00-10:00', '服务器基础', 'Linux命令', '掌握服务器'),
        ('第10周', '💎 高阶篇', 6, '', '周一至周六', '10:30-11:30', 'Nginx配置', '反向代理', '掌握Nginx'),
        ('第10周', '💎 高阶篇', 6, '', '周一至周六', '14:00-15:00', 'PM2进程管理', '进程守护', '掌握PM2'),
        ('第10周', '💎 高阶篇', 6, '', '周一至周六', '15:30-16:30', 'Docker基础', '容器部署', '掌握Docker'),
        ('第10周', '💎 高阶篇', 6, '', '周一至周六', '17:00-18:00', '监控与日志', '监控体系', '掌握监控'),
        ('第11周', '🎯 实战篇', 6, '', '周一至周六', '9:00-10:00', '项目选题', '团队组建', '确定项目'),
        ('第11周', '🎯 实战篇', 6, '', '周一至周六', '10:30-11:30', '系统设计', '架构设计', '完成设计'),
        ('第11周', '🎯 实战篇', 6, '', '周一至周六', '14:00-15:00', '前端开发', '组件开发', '完成前端'),
        ('第11周', '🎯 实战篇', 6, '', '周一至周六', '15:30-16:30', '后端开发', 'API开发', '完成后端'),
        ('第11周', '🎯 实战篇', 6, '', '周一至周六', '17:00-18:00', '测试与部署', '测试部署', '完成上线'),
        ('第12周', '🎯 实战篇', 6, '', '周一至周六', '9:00-10:00', '项目答辩', '项目展示', '毕业答辩'),
        ('第12周', '🎯 实战篇', 6, '', '周一至周六', '10:30-11:30', '课程总结', '知识梳理', '总结复盘'),
        ('第12周', '🎯 实战篇', 6, '', '周一至周六', '14:00-15:00', '简历制作', '技术栈梳理', '完成简历'),
        ('第12周', '🎯 实战篇', 6, '', '周一至周六', '15:30-16:30', '面试指导', '模拟面试', '面试培训'),
        ('第12周', '🎯 实战篇', 6, '', '周一至周六', '17:00-18:00', '结业典礼', '颁发证书', '正式毕业'),
    ]

    for row_num, row_data in enumerate(schedule_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            set_cell_style(cell)
            set_border(cell)

    col_widths = [8, 12, 8, 10, 12, 10, 18, 18, 15]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:I{len(schedule_data)+1}"

    wb.save(os.path.join(OUTPUT_DIR, '01_课程表.xlsx'))
    print(f"✅ 已生成：01_课程表.xlsx")

def create_assessment_excel():
    """创建考核表Excel"""
    wb = Workbook()

    ws = wb.active
    ws.title = "考核标准"

    headers = ['要求项', '标准', '权重', '评分方式', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_header_style(cell)

    data = [
        ('出勤率', '≥ 90%', '10%', '班主任记录', '请假需提供证明'),
        ('作业完成', '全部提交 + 及格以上', '20%', '作业系统评分', '迟交扣50%'),
        ('阶段测试', '每阶段≥60分', '20%', '在线测试系统', '可补考一次'),
        ('实践项目', '完成全部9个实践项目', '30%', '项目评审', '项目答辩'),
        ('结业答辩', '评委会≥60分通过', '20%', '现场答辩', '3人评审团'),
    ]

    for row_num, row_data in enumerate(data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            set_cell_style(cell, align='center')
            set_border(cell)

    for i, width in enumerate([15, 20, 10, 18, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    ws2 = wb.create_sheet("阶段考核")
    headers2 = ['阶段', '周次', '考核内容', '总分', '及格线', '考试形式']
    for col, header in enumerate(headers2, 1):
        cell = ws2.cell(row=1, column=col, value=header)
        set_header_style(cell, '70AD47')

    phase_data = [
        ('🌟 入门篇', '第1-2周', 'HTML/CSS基础测试', '100', '60', '在线测试+实践'),
        ('🔧 基础篇', '第3-5周', 'JavaScript/Node.js测试', '100', '60', '在线测试+项目'),
        ('🚀 进阶篇', '第6-8周', 'React/平台开发测试', '100', '60', '在线测试+组件'),
        ('💎 高阶篇', '第9-10周', 'TS/部署/工程化测试', '100', '60', '在线测试+部署'),
        ('🎯 实战篇', '第11-12周', '综合项目答辩', '100', '60', '现场答辩'),
    ]

    for row_num, row_data in enumerate(phase_data, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws2.cell(row=row_num, column=col_num, value=value)
            set_cell_style(cell, align='center')
            set_border(cell)

    for i, width in enumerate([12, 10, 20, 8, 8, 15], 1):
        ws2.column_dimensions[get_column_letter(i)].width = width

    wb.save(os.path.join(OUTPUT_DIR, '02_考核表.xlsx'))
    print(f"✅ 已生成：02_考核表.xlsx")

def create_project_tracker_excel():
    """创建项目跟踪表Excel"""
    wb = Workbook()

    ws = wb.active
    ws.title = "项目进度跟踪"

    headers = ['项目编号', '项目名称', '负责阶段', '技术栈', '开始日期', '截止日期', '当前状态', '完成度', '负责人', '备注']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_header_style(cell, 'ED7D31')

    projects = [
        ('P1', '个人简历网站', '第一阶段', 'HTML/CSS', '', '', '未开始', '0%', '', ''),
        ('P2', '待办事项应用', '第二阶段', 'JavaScript', '', '', '未开始', '0%', '', ''),
        ('P3', '用户管理系统', '第二阶段', 'Node.js/MySQL', '', '', '未开始', '0%', '', ''),
        ('P4', 'AI对话组件', '第三阶段', 'React', '', '', '未开始', '0%', '', ''),
        ('P5', '会员中心', '第三阶段', 'React/Node.js', '', '', '未开始', '0%', '', ''),
        ('P6', '支付功能', '第三阶段', '支付API', '', '', '未开始', '0%', '', ''),
        ('P7', 'TypeScript重构', '第四阶段', 'TypeScript', '', '', '未开始', '0%', '', ''),
        ('P8', '容器化部署', '第四阶段', 'Docker/Nginx', '', '', '未开始', '0%', '', ''),
        ('P9', '结业综合项目', '第五阶段', '全栈技术', '', '', '未开始', '0%', '', ''),
    ]

    for row_num, row_data in enumerate(projects, 2):
        for col_num, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_num, column=col_num, value=value)
            set_cell_style(cell, align='center')
            set_border(cell)

    for i, width in enumerate([10, 15, 12, 15, 12, 12, 10, 8, 10, 15], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    wb.save(os.path.join(OUTPUT_DIR, '03_项目跟踪表.xlsx'))
    print(f"✅ 已生成：03_项目跟踪表.xlsx")

def create_student_progress_excel():
    """创建学员进度表Excel"""
    wb = Workbook()

    ws = wb.active
    ws.title = "学员学习进度"

    headers = ['学号', '姓名', '第一阶段', '第二阶段', '第三阶段', '第四阶段', '第五阶段', '总进度', '平均分', '排名']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        set_header_style(cell, '7030A0')

    for i, width in enumerate([10, 12, 12, 12, 12, 12, 12, 10, 10, 8], 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    for row in range(2, 32):
        for col in range(1, 11):
            cell = ws.cell(row=row, column=col, value='')
            set_cell_style(cell, align='center')
            set_border(cell)
        ws.cell(row=row, column=1, value=f'S{row-1:03d}')

    wb.save(os.path.join(OUTPUT_DIR, '04_学员进度表.xlsx'))
    print(f"✅ 已生成：04_学员进度表.xlsx")

def main():
    print("=" * 50)
    print("超无穹AI平台培训课程 - Excel文档生成")
    print("=" * 50)

    if not os.path.exists(OUTPUT_DIR):
        os.makedirs(OUTPUT_DIR)

    create_schedule_excel()
    create_assessment_excel()
    create_project_tracker_excel()
    create_student_progress_excel()

    print("=" * 50)
    print(f"所有Excel文档已生成到：{OUTPUT_DIR}")
    print("=" * 50)

if __name__ == '__main__':
    main()
